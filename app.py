import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import subprocess
import shutil
import time
import tempfile
from datetime import datetime

# ================================================
# Page config
# ================================================
st.set_page_config(page_title="Report Generator", page_icon="📊", layout="centered")
st.title("📊 Property Report Generator")
st.markdown("Upload your two files below, then click **Generate Reports**.")

# ================================================
# File uploaders
# ================================================
col1, col2 = st.columns(2)

with col1:
    source_upload = st.file_uploader(
        "📁 Management Fee sheet",
        type=["xlsx", "xls"],
        help="File name must contain 'management fee'"
    )

with col2:
    template_upload = st.file_uploader(
        "📋 Analysis Sheet template",
        type=["xlsx", "xls"],
        help="File name must contain 'analysis sheet'"
    )

# ================================================
# Validate uploads by filename
# ================================================
source_ok = source_upload and "management fee" in source_upload.name.lower()
template_ok = template_upload and "analysis sheet" in template_upload.name.lower()

if source_upload and not source_ok:
    st.error("⚠️ The source file name must contain **'management fee'**.")

if template_upload and not template_ok:
    st.error("⚠️ The template file name must contain **'analysis sheet'**.")

# ================================================
# Generate button
# ================================================
generate_clicked = st.button(
    "⚙️ Generate Reports",
    disabled=not (source_ok and template_ok),
    use_container_width=True,
    type="primary"
)

if generate_clicked:
    # Derive month/year label from today's date at generation time
    now = datetime.now()
    month_label = now.strftime("%B %Y")   # e.g. "May 2026"

    with tempfile.TemporaryDirectory() as tmp_dir:
        # Save uploaded files to temp dir
        source_path = os.path.join(tmp_dir, source_upload.name)
        template_path = os.path.join(tmp_dir, template_upload.name)

        with open(source_path, "wb") as f:
            f.write(source_upload.getbuffer())
        with open(template_path, "wb") as f:
            f.write(template_upload.getbuffer())

        # Output directory inside temp dir
        output_dir = os.path.join(tmp_dir, "Generated_Reports")
        os.makedirs(output_dir, exist_ok=True)

        # ── Load and group data ──────────────────────────────────
        df = pd.read_excel(source_path)
        property_col = df.columns[0]
        grouped = df.groupby(property_col)
        total = grouped.ngroups

        # ── Locate LibreOffice ───────────────────────────────────
        libre_exec = shutil.which("libreoffice") or shutil.which("soffice")
        if not libre_exec:
            st.error(
                "LibreOffice not found on this machine. "
                "Please install it from https://www.libreoffice.org/download/libreoffice/"
            )
            st.stop()

        # ── Progress bar ─────────────────────────────────────────
        progress_bar = st.progress(0, text="Starting…")
        status = st.empty()

        for i, (property_name, group) in enumerate(grouped):
            status.info(f"Processing: **{property_name}** ({i + 1} of {total})")

            wb = load_workbook(template_path)
            ws = wb.active
            start_row = 9

            # Insert grouped rows
            for row_idx, row in enumerate(group.values.tolist(), start=start_row):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            last_row = start_row + len(group) - 1
            sum_row = last_row + 1

            # Format currency columns H–J
            thin_border = Border(bottom=Side(style="thin"))
            gbp_format = '£#,##0.00'

            for col_letter in ["H", "I", "J"]:
                ws[f"{col_letter}{last_row}"].border = thin_border
                for row_idx in range(start_row, last_row + 1):
                    ws[f"{col_letter}{row_idx}"].number_format = gbp_format
                ws[f"{col_letter}{sum_row}"] = f"=SUM({col_letter}{start_row}:{col_letter}{last_row})"
                ws[f"{col_letter}{sum_row}"].number_format = gbp_format

            # Format D and E as short dates with left alignment
            date_format = "DD/MM/YYYY"
            left_align = Alignment(horizontal="left")

            for col_idx in [4, 5]:
                col_letter = get_column_letter(col_idx)
                for row_idx in range(start_row, last_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.number_format = date_format
                    cell.alignment = left_align

            # Auto-adjust column widths
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = max(
                    (len(str(cell.value)) for cell in ws[col_letter] if cell.value), default=0
                )
                ws.column_dimensions[col_letter].width = max_length + 2

            # Save Excel — name uses dynamic month/year
            clean_name = "".join(
                c for c in str(property_name) if c.isalnum() or c in (" ", "-", "_")
            ).strip()
            filename_base = f"{clean_name} - {month_label} analysis sheet"
            xlsx_path = os.path.join(output_dir, f"{filename_base}.xlsx")
            wb.save(xlsx_path)

            # Convert to PDF via LibreOffice
            time.sleep(1)
            subprocess.run([
                libre_exec,
                "--headless",
                "--convert-to", "pdf",
                "--outdir", output_dir,
                xlsx_path
            ], check=True)

            progress_bar.progress((i + 1) / total, text=f"{i + 1} / {total} properties done")

        status.empty()
        progress_bar.empty()

        # ── Create ZIP ───────────────────────────────────────────
        zip_name = f"All_Reports_{now.strftime('%B_%Y')}"   # e.g. All_Reports_May_2026
        zip_base = os.path.join(tmp_dir, zip_name)
        zip_full_path = shutil.make_archive(zip_base, "zip", output_dir)

        # Read ZIP into memory before temp dir is cleaned up
        with open(zip_full_path, "rb") as zf:
            zip_bytes = zf.read()

    # ── Success + download ───────────────────────────────────────
    st.success(f"✅ All {total} reports generated for **{month_label}**!")

    st.download_button(
        label="⬇️ Download All Reports (.zip)",
        data=zip_bytes,
        file_name=f"{zip_name}.zip",
        mime="application/zip",
        use_container_width=True,
    )
